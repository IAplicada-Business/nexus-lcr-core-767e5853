"""
src/sci/gerar_planilha.py

Gera a planilha de importação SCI no formato exato.
11 colunas: DATA | DÉBITO | CRÉDITO | PART DÉB | PART CRED | VALOR |
            HISTÓRICO | COMPLEMENTO | DOCUMENTO | CC DÉB | CC CRED

Baseado no modelo real: 08_-_Modelo_Planilha_Importacao_Lctos_SCI.xls
"""

import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


# Colunas exatas do SCI — ordem e nomes conforme modelo
COLUNAS_SCI = [
    'DATA',
    'DÉBITO',
    'CRÉDITO',
    'PART DÉB,',      # vírgula faz parte do nome original
    'PART, CRED',     # vírgula faz parte do nome original
    'VALOR',
    'HISTÓRICO',
    'COMPLEMENTO',
    'DOCUMENTO',
    'CENTRO DE CUSTO DÉB',
    'CENTRO DE CUSTO CRED'
]


def gerar_planilha_sci(
    linhas: list,
    nome_empresa: str,
    competencia: str,
    caminho_saida: str
) -> str:
    """
    Gera planilha .xls no formato exato de importação do SCI.

    linhas = lista de dicts retornados pelo motor de classificação:
    [{
        'data': 'YYYYMMDD',
        'debito': int,
        'credito': int,
        'part_deb': int ou None,
        'part_cred': int ou None,
        'valor': float,
        'historico': int,
        'complemento': str,
        'documento': str ou None,
        'centro_custo_deb': str ou None,
        'centro_custo_cred': str ou None
    }]

    Retorna o caminho do arquivo gerado.
    """

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Planilha de importação'

    # ── Cabeçalho ──────────────────────────────────────────────────────
    cabecalho_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    cabecalho_font = Font(bold=True, size=10)
    thin = Side(border_style='thin', color='000000')
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, nome_col in enumerate(COLUNAS_SCI, start=1):
        celula = ws.cell(row=1, column=col_idx, value=nome_col)
        celula.fill = cabecalho_fill
        celula.font = cabecalho_font
        celula.border = borda
        celula.alignment = Alignment(horizontal='center')

    # ── Dados ──────────────────────────────────────────────────────────
    for row_idx, linha in enumerate(linhas, start=2):
        # Normaliza DATA: qualquer formato → "YYYYMMDD" como TEXT (string)
        # SCI espera texto, não inteiro — célula numérica é rejeitada
        data_raw = linha.get('data', '')
        try:
            data_val = str(int(str(data_raw).replace('-', ''))).zfill(8)
        except (ValueError, TypeError):
            data_val = str(data_raw)

        def _n(v):
            """None e 0 viram '' para células vazias do SCI."""
            return '' if v is None else v

        valores = [
            data_val,                          # DATA — YYYYMMDD int
            linha.get('debito'),               # DÉBITO
            linha.get('credito'),              # CRÉDITO
            _n(linha.get('part_deb')),         # PART DÉB
            _n(linha.get('part_cred')),        # PART CRED
            linha.get('valor'),                # VALOR
            linha.get('historico'),            # HISTÓRICO
            linha.get('complemento') or '',    # COMPLEMENTO
            _n(linha.get('documento')),        # DOCUMENTO
            _n(linha.get('centro_custo_deb')), # CC DÉB
            _n(linha.get('centro_custo_cred')),# CC CRED
        ]

        for col_idx, valor in enumerate(valores, start=1):
            celula = ws.cell(row=row_idx, column=col_idx, value=valor)
            celula.border = borda
            celula.font = Font(size=10)

            # Alinhamento por tipo de coluna
            if col_idx in [1, 2, 3, 4, 5, 7, 9]:  # Colunas numéricas
                celula.alignment = Alignment(horizontal='center')
            elif col_idx == 6:  # Valor — plain float, sem number_format
                celula.alignment = Alignment(horizontal='right')
            else:
                celula.alignment = Alignment(horizontal='left')

    # ── Larguras das colunas ────────────────────────────────────────────
    larguras = [12, 8, 8, 10, 10, 12, 10, 60, 12, 20, 20]
    for col_idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = largura

    # ── Nome do arquivo ─────────────────────────────────────────────────
    # Formato baseado no modelo real:
    # "08 - Modelo Planilha Importação Lctos.xls"
    mes_ano = competencia.replace('/', '-')
    nome_arquivo = f"{nome_empresa} - Lancamentos {mes_ano}.xlsx"
    caminho_completo = os.path.join(caminho_saida, nome_arquivo)

    os.makedirs(caminho_saida, exist_ok=True)
    wb.save(caminho_completo)

    print(f"Planilha gerada: {caminho_completo}")
    print(f"Total de lançamentos: {len(linhas)}")

    return caminho_completo


def validar_planilha(caminho: str) -> dict:
    """Valida a planilha gerada antes de enviar ao LevelDrive."""
    wb = openpyxl.load_workbook(caminho)
    ws = wb.active

    erros = []
    avisos = []

    # Verifica cabeçalho
    for col_idx, nome_esperado in enumerate(COLUNAS_SCI, start=1):
        valor_celula = ws.cell(row=1, column=col_idx).value
        if valor_celula != nome_esperado:
            erros.append(f"Coluna {col_idx}: esperado '{nome_esperado}', encontrado '{valor_celula}'")

    # Verifica dados
    for row_idx in range(2, ws.max_row + 1):
        data = ws.cell(row=row_idx, column=1).value
        debito = ws.cell(row=row_idx, column=2).value
        credito = ws.cell(row=row_idx, column=3).value
        valor = ws.cell(row=row_idx, column=6).value
        historico = ws.cell(row=row_idx, column=7).value

        if not data:
            erros.append(f"Linha {row_idx}: DATA vazia")
        elif len(str(data)) != 8:
            erros.append(f"Linha {row_idx}: DATA inválida '{data}' (esperado YYYYMMDD)")

        if not debito:
            erros.append(f"Linha {row_idx}: DÉBITO vazio")

        if not credito:
            erros.append(f"Linha {row_idx}: CRÉDITO vazio")

        if not valor or valor == 0:
            avisos.append(f"Linha {row_idx}: VALOR zerado")

        if not historico:
            avisos.append(f"Linha {row_idx}: HISTÓRICO vazio")

    total_linhas = ws.max_row - 1  # Descontando cabeçalho

    return {
        'valida': len(erros) == 0,
        'total_linhas': total_linhas,
        'erros': erros,
        'avisos': avisos
    }


def gerar_revisao_csv(
    itens_revisao: list,
    nome_empresa: str,
    competencia: str,
    caminho_saida: str
) -> str:
    """
    Gera CSV com lançamentos que precisam de revisão manual.
    Inclui colunas extras: sugestão da IA e justificativa.
    """
    mes_ano = competencia.replace('/', '-')
    nome_arquivo = f"{nome_empresa} - Revisao {mes_ano}.csv"
    caminho_completo = os.path.join(caminho_saida, nome_arquivo)

    os.makedirs(caminho_saida, exist_ok=True)

    colunas = [
        'DATA_ORIG', 'DESCRICAO', 'VALOR', 'TIPO',
        'DEB_SUGERIDO', 'CRED_SUGERIDO', 'HISTORICO_SUGERIDO',
        'COMPLEMENTO_SUGERIDO', 'CONFIANCA', 'JUSTIFICATIVA',
        # Colunas em branco para o contador preencher
        'DEB_FINAL', 'CRED_FINAL', 'HISTORICO_FINAL', 'COMPLEMENTO_FINAL',
    ]

    with open(caminho_completo, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=colunas, delimiter=';')
        writer.writeheader()

        for item in itens_revisao:
            t = item.get('transacao_original', {})
            c = item.get('classificacao_sugerida', {})
            writer.writerow({
                'DATA_ORIG': t.get('data', ''),
                'DESCRICAO': t.get('descricao', ''),
                'VALOR': t.get('valor', ''),
                'TIPO': t.get('tipo', ''),
                'DEB_SUGERIDO': c.get('debito', ''),
                'CRED_SUGERIDO': c.get('credito', ''),
                'HISTORICO_SUGERIDO': c.get('historico', ''),
                'COMPLEMENTO_SUGERIDO': c.get('complemento', ''),
                'CONFIANCA': f"{c.get('confianca', 0):.0%}",
                'JUSTIFICATIVA': c.get('justificativa', ''),
                'DEB_FINAL': '',
                'CRED_FINAL': '',
                'HISTORICO_FINAL': '',
                'COMPLEMENTO_FINAL': '',
            })

    print(f"Revisão manual: {caminho_completo} ({len(itens_revisao)} itens)")
    return caminho_completo


# ─────────────────────────────────────────────
# Teste local com dados reais do modelo
# ─────────────────────────────────────────────

if __name__ == '__main__':
    # Replica exata de algumas linhas do modelo SCI que você enviou
    linhas_teste = [
        {'data': 20260401, 'debito': 9, 'credito': 293, 'part_deb': None, 'part_cred': None,
         'valor': 0.25, 'historico': 1961, 'complemento': '04/2026', 'documento': None},
        {'data': 20260401, 'debito': 465, 'credito': 9, 'part_deb': None, 'part_cred': None,
         'valor': 225.12, 'historico': 2216, 'complemento': '04/2026 - Maria De Lourdes Do Nascimento Silva',
         'documento': None},
        {'data': 20260402, 'debito': 762, 'credito': 9, 'part_deb': None, 'part_cred': None,
         'valor': 18.00, 'historico': 1001,
         'complemento': 'despesas com copa e cozinha cfe. Cupom fiscal nº 83883 - Ze Bolacha Com. Generos Alimenticio',
         'documento': None},
        {'data': 20260410, 'debito': 148, 'credito': 9, 'part_deb': 639735, 'part_cred': None,
         'valor': 1621.00, 'historico': 317, 'complemento': '2491 - Contaosiris Contábil Ltda.',
         'documento': None},
    ]

    resultado = gerar_planilha_sci(
        linhas=linhas_teste,
        nome_empresa='KIALO CONSULTORIA E ENGENHARIA',
        competencia='04/2026',
        caminho_saida='/tmp/lcr-test'
    )

    validacao = validar_planilha(resultado)
    print(f"\n=== Validação ===")
    print(f"Válida: {validacao['valida']}")
    print(f"Linhas: {validacao['total_linhas']}")
    if validacao['erros']:
        print(f"Erros: {validacao['erros']}")
    if validacao['avisos']:
        print(f"Avisos: {validacao['avisos']}")
